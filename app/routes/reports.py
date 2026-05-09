import os
import re
from datetime import date, datetime, timedelta
from io import BytesIO

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import case, desc, func

from app.models.accounting import (
    ClosedWorkerAccount,
    ExpenseCategory,
    EXPENSE_TRANSACTION_TYPE_ALIASES,
    INCOME_TRANSACTION_TYPE_ALIASES,
    Transaction,
    WORKER_REFERENCE_TYPE_ALIASES,
    is_expense_transaction,
    is_income_transaction,
    normalize_reference_type,
    normalize_transaction_type,
)
from app.models.box import BoxPurchase, BoxType, BoxUsage
from app.models.crop import Crop, CropConsumption, Production, Sales
from app.models.inventory import (
    GeneralConsumption,
    InventoryItem,
    InventoryPurchase,
    InventoryTransaction,
)
from app.models.motor import Motor, MotorCost, MotorUsage, OperatorQuota
from app.models.worker import Attendance, MotorLog, WorkLog, Worker

bp = Blueprint("reports", __name__, url_prefix="/reports")

_PDF_FONT_NAME = None
_ARABIC_REGEX = re.compile(r"[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]")
_MOJIBAKE_HINTS = ("Ø", "Ù", "Ã", "â")

try:
    import arabic_reshaper
except Exception:
    arabic_reshaper = None

try:
    from bidi.algorithm import get_display
except Exception:
    get_display = None


def _require_reports_access():
    if not current_user.can_manage_reports and not current_user.is_admin:
        flash("ليس لديك صلاحية للوصول إلى قسم التقارير", "danger")
        return redirect(url_for("home.index"))
    return None


def _require_financial_access():
    if not (
        current_user.can_manage_reports
        or current_user.can_manage_accounting
        or current_user.is_admin
    ):
        flash("ليس لديك صلاحية للوصول إلى هذا التقرير", "danger")
        return redirect(url_for("home.index"))
    return None


def _repair_mojibake_text(value):
    """Best-effort fix for UTF-8 text that was decoded as latin-1/cp1252."""
    if not isinstance(value, str) or not value:
        return value
    if not any(marker in value for marker in _MOJIBAKE_HINTS):
        return value
    try:
        return value.encode("latin-1").decode("utf-8")
    except UnicodeError:
        return value


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _read_date_range():
    from_raw = (request.args.get("from_date") or "").strip()
    to_raw = (request.args.get("to_date") or "").strip()

    from_date = _parse_date(from_raw)
    to_date = _parse_date(to_raw)

    if from_raw and not from_date:
        flash("تنسيق تاريخ البداية غير صحيح", "warning")
    if to_raw and not to_date:
        flash("تنسيق تاريخ النهاية غير صحيح", "warning")

    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
        flash("تم تبديل التاريخين لأن البداية كانت أكبر من النهاية", "info")

    from_str = from_date.strftime("%Y-%m-%d") if from_date else ""
    to_str = to_date.strftime("%Y-%m-%d") if to_date else ""
    return from_date, to_date, from_str, to_str


def _date_range_label(from_date, to_date):
    if from_date and to_date:
        return f"{from_date} -> {to_date}"
    if from_date:
        return f"من {from_date}"
    if to_date:
        return f"حتى {to_date}"
    return "كل الفترات"


def _apply_date_range(query, column, from_date, to_date):
    if from_date:
        query = query.filter(column >= from_date)
    if to_date:
        query = query.filter(column <= to_date)
    return query


def _apply_datetime_range(query, column, from_date, to_date):
    if from_date:
        query = query.filter(column >= datetime.combine(from_date, datetime.min.time()))
    if to_date:
        next_day = to_date + timedelta(days=1)
        query = query.filter(column < datetime.combine(next_day, datetime.min.time()))
    return query


def _query_value(query, aggregate_expression):
    return query.with_entities(func.coalesce(aggregate_expression, 0)).scalar() or 0


def _detect_worker_payment_kind(transaction):
    """Detect worker payment kind from marker/description."""
    notes = transaction.notes or ""
    description = _repair_mojibake_text(transaction.description or "")

    if "worker_payment_kind=loan" in notes:
        return "loan"
    if "worker_payment_kind=advance" in notes:
        return "advance"

    if "سلفة" in description:
        return "loan"
    if "دفعة" in description or "على الحساب" in description:
        return "advance"
    return "other"


def _export_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return round(value, 2)
    if isinstance(value, str):
        return _repair_mojibake_text(value)
    return value


def _shape_text_for_pdf(value):
    """Prepare Arabic text for PDF rendering (reshaping + bidi)."""
    raw_text = _repair_mojibake_text("" if value is None else str(value))
    if not raw_text:
        return raw_text

    if not _ARABIC_REGEX.search(raw_text):
        return raw_text

    if not arabic_reshaper or not get_display:
        return raw_text

    try:
        return get_display(arabic_reshaper.reshape(raw_text))
    except Exception:
        return raw_text


def _pdf_font_name():
    global _PDF_FONT_NAME
    if _PDF_FONT_NAME:
        return _PDF_FONT_NAME

    fallback = "Helvetica"
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception:
        _PDF_FONT_NAME = fallback
        return _PDF_FONT_NAME

    candidates = [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\Tahoma.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for font_path in candidates:
        if not os.path.exists(font_path):
            continue
        try:
            pdfmetrics.registerFont(TTFont("ReportFont", font_path))
            _PDF_FONT_NAME = "ReportFont"
            return _PDF_FONT_NAME
        except Exception:
            continue

    _PDF_FONT_NAME = fallback
    return _PDF_FONT_NAME


def _build_excel_response(report_key, report_title, sections, from_date, to_date):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    sheet.sheet_view.rightToLeft = True

    title_font = Font(name="Arial", bold=True, size=14)
    header_font = Font(name="Arial", bold=True)
    section_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial", size=11)
    section_fill = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
    header_fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")

    row_index = 1
    max_cols = 2

    sheet.cell(row=row_index, column=1, value=report_title).font = title_font
    sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="right")
    row_index += 1

    sheet.cell(row=row_index, column=1, value="الفترة الزمنية").font = header_font
    sheet.cell(row=row_index, column=2, value=_date_range_label(from_date, to_date)).font = body_font
    sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="right")
    sheet.cell(row=row_index, column=2).alignment = Alignment(horizontal="right")
    row_index += 2

    for section in sections:
        title = section.get("title", "Section")
        headers = section.get("headers", [])
        rows = section.get("rows", [])
        header_count = max(1, len(headers))
        max_cols = max(max_cols, header_count)

        sheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=header_count,
        )
        title_cell = sheet.cell(row=row_index, column=1, value=title)
        title_cell.font = section_font
        title_cell.fill = section_fill
        title_cell.alignment = Alignment(horizontal="right")
        row_index += 1

        if headers:
            for col_index, header in enumerate(headers, start=1):
                header_cell = sheet.cell(row=row_index, column=col_index, value=header)
                header_cell.font = header_font
                header_cell.fill = header_fill
                header_cell.alignment = Alignment(horizontal="right")
            row_index += 1

        if not rows:
            placeholder_count = len(headers) if headers else 1
            rows = [["-"] * placeholder_count]

        for row in rows:
            max_cols = max(max_cols, len(row))
            for col_index, value in enumerate(row, start=1):
                row_cell = sheet.cell(
                    row=row_index,
                    column=col_index,
                    value=_export_cell_value(value),
                )
                row_cell.font = body_font
                row_cell.alignment = Alignment(horizontal="right")
            row_index += 1

        row_index += 1

    for col_index in range(1, max_cols + 1):
        col_letter = get_column_letter(col_index)
        max_length = 0
        for cell in sheet[col_letter]:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        sheet.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{report_key}_{stamp}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _build_pdf_response(report_key, report_title, sections, from_date, to_date):
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception:
        flash("تصدير PDF غير متاح حاليًا", "warning")
        return redirect(
            url_for(
                request.endpoint,
                from_date=from_date.strftime("%Y-%m-%d") if from_date else None,
                to_date=to_date.strftime("%Y-%m-%d") if to_date else None,
            )
        )

    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=18, rightMargin=18)
    font_name = _pdf_font_name()

    styles = getSampleStyleSheet()
    title_style = styles["Heading2"].clone("ReportTitle")
    title_style.fontName = font_name
    title_style.fontSize = 14
    title_style.alignment = TA_RIGHT
    title_style.wordWrap = "RTL"

    meta_style = styles["Normal"].clone("ReportMeta")
    meta_style.fontName = font_name
    meta_style.fontSize = 10
    meta_style.alignment = TA_RIGHT
    meta_style.wordWrap = "RTL"

    section_style = styles["Heading4"].clone("ReportSection")
    section_style.fontName = font_name
    section_style.fontSize = 11
    section_style.alignment = TA_RIGHT
    section_style.wordWrap = "RTL"

    flow = [
        Paragraph(_shape_text_for_pdf(report_title), title_style),
        Paragraph(_shape_text_for_pdf(f"الفترة: {_date_range_label(from_date, to_date)}"), meta_style),
        Spacer(1, 10),
    ]

    for section in sections:
        title = section.get("title", "Section")
        headers = section.get("headers", [])
        rows = section.get("rows", [])
        flow.append(Paragraph(_shape_text_for_pdf(title), section_style))

        if not rows:
            placeholder_count = len(headers) if headers else 1
            rows = [["-"] * placeholder_count]

        table_data = []
        if headers:
            table_data.append([_shape_text_for_pdf(header) for header in headers])
        for row in rows:
            table_data.append(
                [_shape_text_for_pdf(_export_cell_value(value)) for value in row]
            )

        table = Table(table_data, repeatRows=1 if headers else 0)
        table_style_data = [
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ]
        if headers:
            table_style_data.extend(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("FONTSIZE", (0, 0), (-1, 0), 8.5),
                ]
            )
        table.setStyle(TableStyle(table_style_data))

        flow.append(table)
        flow.append(Spacer(1, 10))

    document.build(flow)
    output.seek(0)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{report_key}_{stamp}.pdf"
    return send_file(output, mimetype="application/pdf", as_attachment=True, download_name=filename)


def _maybe_export(report_key, report_title, sections, from_date, to_date):
    export_type = (request.args.get("export") or "").strip().lower()
    if export_type == "excel":
        return _build_excel_response(report_key, report_title, sections, from_date, to_date)
    if export_type == "pdf":
        return _build_pdf_response(report_key, report_title, sections, from_date, to_date)
    return None


@bp.route("/")
@login_required
def index():
    """Reports landing page."""
    denied = _require_reports_access()
    if denied:
        return denied

    summary = {
        "workers_count": Worker.query.count(),
        "inventory_items_count": InventoryItem.query.count(),
        "crops_count": Crop.query.count(),
        "sales_count": Sales.query.count(),
        "transactions_count": Transaction.query.count(),
        "motors_count": Motor.query.count(),
    }
    return render_template("reports/index.html", summary=summary)


@bp.route("/workers-detailed")
@login_required
def workers_detailed_report():
    """Detailed workers report."""
    denied = _require_reports_access()
    if denied:
        return denied

    from_date, to_date, from_date_str, to_date_str = _read_date_range()

    selected_worker_id_raw = (request.args.get("worker_id") or "").strip()
    workers = Worker.query.order_by(Worker.is_active.desc(), Worker.name.asc()).all()
    worker_lookup = {worker.id: worker for worker in workers}

    selected_worker_id = None
    if selected_worker_id_raw:
        try:
            candidate_worker_id = int(selected_worker_id_raw)
            if candidate_worker_id in worker_lookup:
                selected_worker_id = candidate_worker_id
            else:
                flash("العامل المحدد غير موجود", "warning")
        except ValueError:
            flash("فلتر العامل غير صالح", "warning")

    selected_worker = worker_lookup.get(selected_worker_id)
    filtered_workers = [selected_worker] if selected_worker else workers
    filtered_worker_ids = {worker.id for worker in filtered_workers}

    active_workers = sum(1 for worker in filtered_workers if worker.is_active)
    inactive_workers = len(filtered_workers) - active_workers
    monthly_workers = sum(1 for worker in filtered_workers if worker.is_monthly)
    hourly_workers = len(filtered_workers) - monthly_workers

    work_logs_query = WorkLog.query
    motor_logs_query = MotorLog.query
    attendance_query = Attendance.query
    closed_accounts_query = ClosedWorkerAccount.query

    if selected_worker_id is not None:
        work_logs_query = work_logs_query.filter(WorkLog.worker_id == selected_worker_id)
        motor_logs_query = motor_logs_query.filter(MotorLog.worker_id == selected_worker_id)
        attendance_query = attendance_query.filter(Attendance.worker_id == selected_worker_id)
        closed_accounts_query = closed_accounts_query.filter(
            ClosedWorkerAccount.worker_id == selected_worker_id
        )

    work_logs_query = _apply_date_range(work_logs_query, WorkLog.work_date, from_date, to_date)
    motor_logs_query = _apply_datetime_range(
        motor_logs_query, MotorLog.start_date, from_date, to_date
    )
    attendance_query = _apply_date_range(
        attendance_query, Attendance.attendance_date, from_date, to_date
    )
    closed_accounts_query = _apply_date_range(
        closed_accounts_query,
        ClosedWorkerAccount.closure_date,
        from_date,
        to_date,
    )

    work_log_count = work_logs_query.count()
    attendance_hours = _query_value(
        attendance_query.filter(
            (Attendance.status == "حاضر") | (Attendance.is_present.is_(True))
        ),
        func.sum(Attendance.hours_worked),
    )

    if work_log_count:
        total_work_hours = _query_value(work_logs_query, func.sum(WorkLog.hours))
        worker_hours_rows = (
            work_logs_query.with_entities(
                WorkLog.worker_id.label("worker_id"),
                func.coalesce(func.sum(WorkLog.hours), 0).label("hours"),
                func.count(WorkLog.id).label("entries"),
            )
            .group_by(WorkLog.worker_id)
            .all()
        )
        worker_days_rows = (
            work_logs_query.with_entities(
                WorkLog.worker_id.label("worker_id"),
                func.count(func.distinct(WorkLog.work_date)).label("worked_days"),
            )
            .group_by(WorkLog.worker_id)
            .all()
        )
        top_workers_by_hours = (
            work_logs_query.join(Worker, Worker.id == WorkLog.worker_id)
            .with_entities(
                Worker.id,
                Worker.name,
                func.coalesce(func.sum(WorkLog.hours), 0).label("hours"),
                func.count(WorkLog.id).label("entries"),
            )
            .group_by(Worker.id, Worker.name)
            .order_by(desc("hours"))
            .limit(15)
            .all()
        )
    else:
        total_work_hours = attendance_hours
        attended_query = attendance_query.filter(
            (Attendance.status == "حاضر") | (Attendance.is_present.is_(True))
        )
        worker_hours_rows = (
            attended_query.with_entities(
                Attendance.worker_id.label("worker_id"),
                func.coalesce(func.sum(Attendance.hours_worked), 0).label("hours"),
                func.count(Attendance.id).label("entries"),
            )
            .group_by(Attendance.worker_id)
            .all()
        )
        worker_days_rows = (
            attended_query.with_entities(
                Attendance.worker_id.label("worker_id"),
                func.count(func.distinct(Attendance.attendance_date)).label("worked_days"),
            )
            .group_by(Attendance.worker_id)
            .all()
        )
        top_workers_by_hours = (
            attended_query.join(Worker, Worker.id == Attendance.worker_id)
            .with_entities(
                Worker.id,
                Worker.name,
                func.coalesce(func.sum(Attendance.hours_worked), 0).label("hours"),
                func.count(Attendance.id).label("entries"),
            )
            .group_by(Worker.id, Worker.name)
            .order_by(desc("hours"))
            .limit(15)
            .all()
        )

    total_motor_hours = _query_value(motor_logs_query, func.sum(MotorLog.total_hours))
    attendance_count = attendance_query.count()

    attendance_status_rows = (
        attendance_query.with_entities(Attendance.status, func.count(Attendance.id))
        .group_by(Attendance.status)
        .all()
    )
    attendance_status = {
        (status or "غير محدد"): count for status, count in attendance_status_rows
    }

    hours_by_worker = {
        row.worker_id: float(row.hours or 0.0) for row in worker_hours_rows
    }
    entries_by_worker = {
        row.worker_id: int(row.entries or 0) for row in worker_hours_rows
    }
    worked_days_by_worker = {
        row.worker_id: int(row.worked_days or 0) for row in worker_days_rows
    }

    worker_transactions_query = Transaction.query.filter(
        Transaction.reference_type.in_(WORKER_REFERENCE_TYPE_ALIASES)
    )
    if selected_worker_id is not None:
        worker_transactions_query = worker_transactions_query.filter(
            Transaction.reference_id == selected_worker_id
        )
    worker_transactions_query = _apply_date_range(
        worker_transactions_query,
        Transaction.transaction_date,
        from_date,
        to_date,
    )
    worker_transactions = worker_transactions_query.all()

    worker_payment_summary = {}
    for transaction in worker_transactions:
        reference_id = transaction.reference_id
        if reference_id is None or reference_id not in filtered_worker_ids:
            continue
        if not is_expense_transaction(transaction.transaction_type):
            continue

        amount = float(transaction.amount or 0.0)
        summary = worker_payment_summary.setdefault(
            reference_id,
            {"loans": 0.0, "advances": 0.0, "total_payments": 0.0},
        )
        payment_kind = _detect_worker_payment_kind(transaction)
        if payment_kind == "loan":
            summary["loans"] += amount
        elif payment_kind == "advance":
            summary["advances"] += amount
        summary["total_payments"] += amount

    worker_financial_rows = []
    for worker in filtered_workers:
        total_hours_for_worker = float(hours_by_worker.get(worker.id, 0.0))
        worked_days_for_worker = int(worked_days_by_worker.get(worker.id, 0))
        total_entries_for_worker = int(entries_by_worker.get(worker.id, 0))

        calculated_salary = (
            float(worker.monthly_salary or 0.0)
            if worker.is_monthly
            else total_hours_for_worker * float(worker.hourly_rate or 0.0)
        )

        payments = worker_payment_summary.get(
            worker.id,
            {"loans": 0.0, "advances": 0.0, "total_payments": 0.0},
        )
        total_loans = float(payments["loans"])
        total_advances = float(payments["advances"])
        total_payments = float(payments["total_payments"])
        remaining_salary = calculated_salary - total_payments

        worker_financial_rows.append(
            {
                "worker_id": worker.id,
                "worker_name": worker.name,
                "worker_type": "شهري" if worker.is_monthly else "بالساعة",
                "total_hours": total_hours_for_worker,
                "worked_days": worked_days_for_worker,
                "entries": total_entries_for_worker,
                "salary": calculated_salary,
                "loans": total_loans,
                "advances": total_advances,
                "total_payments": total_payments,
                "remaining_salary": remaining_salary,
            }
        )

    total_calculated_salaries = sum(row["salary"] for row in worker_financial_rows)
    total_worker_loans = sum(row["loans"] for row in worker_financial_rows)
    total_worker_advances = sum(row["advances"] for row in worker_financial_rows)
    total_worker_payments = sum(row["total_payments"] for row in worker_financial_rows)
    total_remaining_balances = sum(row["remaining_salary"] for row in worker_financial_rows)

    closed_accounts = closed_accounts_query.order_by(ClosedWorkerAccount.closure_date.desc()).all()
    total_closed_balance = sum(account.final_balance or 0 for account in closed_accounts)

    worker_scope_label = f" - {selected_worker.name}" if selected_worker else ""

    export_sections = [
        {
            "title": "ملخص العمال",
            "headers": ["المؤشر", "القيمة"],
            "rows": [
                ["إجمالي العمال في النتيجة", len(filtered_workers)],
                ["العمال النشطون", active_workers],
                ["العمال المعطلون", inactive_workers],
                ["عمال شهري", monthly_workers],
                ["عمال بالساعة", hourly_workers],
                ["سجلات الحضور", attendance_count],
                ["إجمالي ساعات العمل", total_work_hours],
                ["إجمالي ساعات المحركات", total_motor_hours],
                ["إجمالي الرواتب المحسوبة", total_calculated_salaries],
                ["إجمالي السلف", total_worker_loans],
                ["إجمالي دفعات الحساب", total_worker_advances],
                ["إجمالي الحسومات على العمال", total_worker_payments],
                ["إجمالي المتبقي", total_remaining_balances],
                ["عدد الحسابات المسكرة", len(closed_accounts)],
                ["إجمالي أرصدة الحسابات المسكرة", total_closed_balance],
            ],
        },
        {
            "title": "ملخص مالي وتشغيلي لكل عامل",
            "headers": [
                "العامل",
                "النوع",
                "إجمالي الساعات",
                "أيام العمل",
                "عدد السجلات",
                "الراتب المحسوب",
                "السلف",
                "دفعات الحساب",
                "إجمالي الحسومات",
                "المتبقي",
            ],
            "rows": [
                [
                    row["worker_name"],
                    row["worker_type"],
                    row["total_hours"],
                    row["worked_days"],
                    row["entries"],
                    row["salary"],
                    row["loans"],
                    row["advances"],
                    row["total_payments"],
                    row["remaining_salary"],
                ]
                for row in worker_financial_rows
            ],
        },
        {
            "title": "حالة الحضور",
            "headers": ["الحالة", "العدد"],
            "rows": [[status, count] for status, count in attendance_status.items()],
        },
        {
            "title": "أفضل العمال حسب ساعات العمل",
            "headers": ["العامل", "إجمالي الساعات", "عدد السجلات"],
            "rows": [
                [row.name, row.hours, row.entries]
                for row in top_workers_by_hours
            ],
        },
        {
            "title": "الحسابات المسكرة",
            "headers": ["العامل", "تاريخ التسكير", "الرصيد النهائي", "السبب"],
            "rows": [
                [
                    account.worker_name,
                    account.closure_date,
                    account.final_balance or 0,
                    account.closure_reason or "-",
                ]
                for account in closed_accounts
            ],
        },
    ]
    export_response = _maybe_export(
        "workers_detailed_report",
        f"تقرير العمال التفصيلي{worker_scope_label}",
        export_sections,
        from_date,
        to_date,
    )
    if export_response:
        return export_response

    return render_template(
        "reports/workers_detailed_report.html",
        workers=workers,
        filtered_workers=filtered_workers,
        selected_worker=selected_worker,
        selected_worker_id=selected_worker_id,
        active_workers=active_workers,
        inactive_workers=inactive_workers,
        monthly_workers=monthly_workers,
        hourly_workers=hourly_workers,
        total_work_hours=total_work_hours,
        total_motor_hours=total_motor_hours,
        attendance_count=attendance_count,
        attendance_status=attendance_status,
        worker_financial_rows=worker_financial_rows,
        total_calculated_salaries=total_calculated_salaries,
        total_worker_loans=total_worker_loans,
        total_worker_advances=total_worker_advances,
        total_worker_payments=total_worker_payments,
        total_remaining_balances=total_remaining_balances,
        top_workers_by_hours=top_workers_by_hours,
        closed_accounts=closed_accounts,
        total_closed_balance=total_closed_balance,
        from_date=from_date_str,
        to_date=to_date_str,
        date_range_label=_date_range_label(from_date, to_date),
    )


@bp.route("/inventory-detailed")
@login_required
def inventory_detailed_report():
    """Detailed inventory report."""
    denied = _require_reports_access()
    if denied:
        return denied

    from_date, to_date, from_date_str, to_date_str = _read_date_range()

    items = InventoryItem.query.order_by(InventoryItem.category.asc(), InventoryItem.name.asc()).all()
    out_of_stock_items = [item for item in items if (item.quantity or 0) <= 0]
    low_stock_items = [item for item in items if 0 < (item.quantity or 0) <= 10]

    stock_value = sum((item.quantity or 0) * (item.purchase_price or 0) for item in items)

    purchases_query = _apply_date_range(
        InventoryPurchase.query, InventoryPurchase.purchase_date, from_date, to_date
    )
    purchase_total_cost = _query_value(purchases_query, func.sum(InventoryPurchase.total_cost))
    purchase_total_qty = _query_value(purchases_query, func.sum(InventoryPurchase.quantity))

    transactions_query = _apply_datetime_range(
        InventoryTransaction.query, InventoryTransaction.transaction_date, from_date, to_date
    )
    in_qty = _query_value(
        transactions_query.filter(InventoryTransaction.transaction_type == "Ø¯Ø®ÙˆÙ„"),
        func.sum(InventoryTransaction.quantity),
    )
    out_qty = _query_value(
        transactions_query.filter(InventoryTransaction.transaction_type == "Ø®Ø±ÙˆØ¬"),
        func.sum(InventoryTransaction.quantity),
    )

    category_summary = (
        InventoryItem.query.with_entities(
            InventoryItem.category,
            func.count(InventoryItem.id).label("items_count"),
            func.coalesce(func.sum(InventoryItem.quantity), 0).label("total_qty"),
        )
        .group_by(InventoryItem.category)
        .order_by(InventoryItem.category.asc())
        .all()
    )

    general_consumptions_query = _apply_date_range(
        GeneralConsumption.query, GeneralConsumption.consumption_date, from_date, to_date
    )
    top_consumed_items = (
        general_consumptions_query.join(
            InventoryItem, InventoryItem.id == GeneralConsumption.inventory_item_id
        )
        .with_entities(
            InventoryItem.name.label("item_name"),
            InventoryItem.category.label("category"),
            func.coalesce(func.sum(GeneralConsumption.quantity_used), 0).label("used_qty"),
            func.count(GeneralConsumption.id).label("entries"),
        )
        .group_by(InventoryItem.id, InventoryItem.name, InventoryItem.category)
        .order_by(desc("used_qty"))
        .limit(15)
        .all()
    )

    recent_purchases = purchases_query.order_by(InventoryPurchase.purchase_date.desc()).limit(30).all()
    recent_transactions = (
        transactions_query.order_by(InventoryTransaction.transaction_date.desc()).limit(40).all()
    )
    recent_general_consumptions = (
        general_consumptions_query.order_by(GeneralConsumption.consumption_date.desc()).limit(30).all()
    )

    box_types_count = BoxType.query.count()
    box_purchases_query = _apply_date_range(BoxPurchase.query, BoxPurchase.purchase_date, from_date, to_date)
    box_usage_query = _apply_date_range(BoxUsage.query, BoxUsage.usage_date, from_date, to_date)
    box_purchases_total_qty = _query_value(box_purchases_query, func.sum(BoxPurchase.quantity))
    box_purchases_total_cost = _query_value(box_purchases_query, func.sum(BoxPurchase.total_cost))
    box_usage_total_qty = _query_value(box_usage_query, func.sum(BoxUsage.quantity_used))
    box_usage_total_cost = _query_value(box_usage_query, func.sum(BoxUsage.total_cost))

    export_sections = [
        {
            "title": "Ù…Ù„Ø®Øµ Ø§Ù„Ù…Ø®Ø²ÙˆÙ†",
            "headers": ["Ø§Ù„Ù…Ø¤Ø´Ø±", "Ø§Ù„Ù‚ÙŠÙ…Ø©"],
            "rows": [
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ø¹Ù†Ø§ØµØ±", len(items)],
                ["Ø¹Ù†Ø§ØµØ± Ù†ÙØ¯ Ù…Ø®Ø²ÙˆÙ†Ù‡Ø§", len(out_of_stock_items)],
                ["Ø¹Ù†Ø§ØµØ± Ù…Ù†Ø®ÙØ¶Ø© Ø§Ù„Ù…Ø®Ø²ÙˆÙ†", len(low_stock_items)],
                ["Ù‚ÙŠÙ…Ø© Ø§Ù„Ù…Ø®Ø²ÙˆÙ†", stock_value],
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ ÙƒÙ…ÙŠØ© Ø§Ù„Ø´Ø±Ø§Ø¡ (Ø¶Ù…Ù† Ø§Ù„ÙÙ„ØªØ±Ø©)", purchase_total_qty],
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ ØªÙƒÙ„ÙØ© Ø§Ù„Ø´Ø±Ø§Ø¡ (Ø¶Ù…Ù† Ø§Ù„ÙÙ„ØªØ±Ø©)", purchase_total_cost],
                ["ÙƒÙ…ÙŠØ© Ø¯Ø®ÙˆÙ„ (Ø¶Ù…Ù† Ø§Ù„ÙÙ„ØªØ±Ø©)", in_qty],
                ["ÙƒÙ…ÙŠØ© Ø®Ø±ÙˆØ¬ (Ø¶Ù…Ù† Ø§Ù„ÙÙ„ØªØ±Ø©)", out_qty],
                ["Ø¹Ø¯Ø¯ Ø£Ù†ÙˆØ§Ø¹ Ø§Ù„ØµÙ†Ø§Ø¯ÙŠÙ‚", box_types_count],
                ["Ù…Ø´ØªØ±ÙŠØ§Øª Ø§Ù„ØµÙ†Ø§Ø¯ÙŠÙ‚ ÙƒÙ…ÙŠØ©", box_purchases_total_qty],
                ["Ù…Ø´ØªØ±ÙŠØ§Øª Ø§Ù„ØµÙ†Ø§Ø¯ÙŠÙ‚ ØªÙƒÙ„ÙØ©", box_purchases_total_cost],
                ["Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø§Ù„ØµÙ†Ø§Ø¯ÙŠÙ‚ ÙƒÙ…ÙŠØ©", box_usage_total_qty],
                ["Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø§Ù„ØµÙ†Ø§Ø¯ÙŠÙ‚ ØªÙƒÙ„ÙØ©", box_usage_total_cost],
            ],
        },
        {
            "title": "Ù…Ù„Ø®Øµ Ø­Ø³Ø¨ Ø§Ù„ÙØ¦Ø©",
            "headers": ["Ø§Ù„ÙØ¦Ø©", "Ø¹Ø¯Ø¯ Ø§Ù„Ø¹Ù†Ø§ØµØ±", "Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„ÙƒÙ…ÙŠØ©"],
            "rows": [[row.category, row.items_count, row.total_qty] for row in category_summary],
        },
        {
            "title": "Ø£ÙƒØ«Ø± Ø§Ù„Ù…ÙˆØ§Ø¯ Ø§Ø³ØªÙ‡Ù„Ø§ÙƒØ§Ù‹",
            "headers": ["Ø§Ù„Ù…Ø§Ø¯Ø©", "Ø§Ù„ÙØ¦Ø©", "Ø§Ù„ÙƒÙ…ÙŠØ©", "Ø¹Ø¯Ø¯ Ø§Ù„Ø³Ø¬Ù„Ø§Øª"],
            "rows": [
                [row.item_name, row.category, row.used_qty, row.entries]
                for row in top_consumed_items
            ],
        },
        {
            "title": "Ø¢Ø®Ø± Ù…Ø´ØªØ±ÙŠØ§Øª Ø§Ù„Ù…Ø®Ø²ÙˆÙ†",
            "headers": ["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„Ø¹Ù†ØµØ±", "Ø§Ù„ÙƒÙ…ÙŠØ©", "Ø§Ù„ØªÙƒÙ„ÙØ©"],
            "rows": [
                [purchase.purchase_date, purchase.item.name, purchase.quantity, purchase.total_cost]
                for purchase in recent_purchases
            ],
        },
        {
            "title": "Ø¢Ø®Ø± Ø­Ø±ÙƒØ§Øª Ø§Ù„Ù…Ø®Ø²ÙˆÙ†",
            "headers": ["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„Ø¹Ù†ØµØ±", "Ø§Ù„Ù†ÙˆØ¹", "Ø§Ù„ÙƒÙ…ÙŠØ©"],
            "rows": [
                [
                    transaction.transaction_date,
                    transaction.item.name,
                    transaction.transaction_type,
                    transaction.quantity,
                ]
                for transaction in recent_transactions
            ],
        },
        {
            "title": "Ø¢Ø®Ø± Ø³Ø¬Ù„Ø§Øª Ø§Ù„Ø§Ø³ØªÙ‡Ù„Ø§Ùƒ Ø§Ù„Ø¹Ø§Ù…",
            "headers": ["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„Ù…Ø§Ø¯Ø©", "Ø§Ù„Ù†ÙˆØ¹", "Ø§Ù„ÙƒÙ…ÙŠØ©"],
            "rows": [
                [
                    consumption.consumption_date,
                    consumption.inventory_item.name,
                    consumption.consumption_type,
                    consumption.quantity_used,
                ]
                for consumption in recent_general_consumptions
            ],
        },
    ]
    export_response = _maybe_export(
        "inventory_detailed_report",
        "ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„ØªÙØµÙŠÙ„ÙŠ",
        export_sections,
        from_date,
        to_date,
    )
    if export_response:
        return export_response

    return render_template(
        "reports/inventory_detailed_report.html",
        items=items,
        out_of_stock_items=out_of_stock_items,
        low_stock_items=low_stock_items,
        stock_value=stock_value,
        purchase_total_cost=purchase_total_cost,
        purchase_total_qty=purchase_total_qty,
        in_qty=in_qty,
        out_qty=out_qty,
        category_summary=category_summary,
        top_consumed_items=top_consumed_items,
        recent_purchases=recent_purchases,
        recent_transactions=recent_transactions,
        recent_general_consumptions=recent_general_consumptions,
        box_types_count=box_types_count,
        box_purchases_total_qty=box_purchases_total_qty,
        box_purchases_total_cost=box_purchases_total_cost,
        box_usage_total_qty=box_usage_total_qty,
        box_usage_total_cost=box_usage_total_cost,
        from_date=from_date_str,
        to_date=to_date_str,
        date_range_label=_date_range_label(from_date, to_date),
    )


@bp.route("/production-detailed")
@login_required
def production_detailed_report():
    """Detailed production report."""
    denied = _require_reports_access()
    if denied:
        return denied

    from_date, to_date, from_date_str, to_date_str = _read_date_range()

    crops = Crop.query.order_by(Crop.is_active.desc(), Crop.name.asc()).all()
    active_crops = sum(1 for crop in crops if crop.is_active)

    production_query = _apply_date_range(Production.query, Production.production_date, from_date, to_date)
    consumptions_query = _apply_date_range(
        CropConsumption.query, CropConsumption.consumption_date, from_date, to_date
    )
    sales_query = _apply_date_range(Sales.query, Sales.sale_date, from_date, to_date)

    total_production_qty = _query_value(production_query, func.sum(Production.quantity))
    production_records_count = production_query.count()
    crop_consumption_qty = _query_value(consumptions_query, func.sum(CropConsumption.quantity_used))
    crop_consumption_count = consumptions_query.count()

    production_by_crop = (
        production_query.join(Crop, Crop.id == Production.crop_id)
        .with_entities(
            Crop.id,
            Crop.name,
            Crop.category,
            func.coalesce(func.sum(Production.quantity), 0).label("qty"),
            func.count(Production.id).label("entries"),
        )
        .group_by(Crop.id, Crop.name, Crop.category)
        .order_by(desc("qty"))
        .all()
    )

    sales_by_crop = (
        sales_query.join(Crop, Crop.id == Sales.crop_id)
        .with_entities(
            Crop.id,
            Crop.name,
            func.coalesce(func.sum(Sales.total_price), 0).label("revenue"),
            func.coalesce(func.sum(Sales.quantity), 0).label("sold_qty"),
        )
        .group_by(Crop.id, Crop.name)
        .order_by(desc("revenue"))
        .all()
    )
    sales_by_crop_map = {row.id: row for row in sales_by_crop}

    crop_performance = []
    for row in production_by_crop:
        sales_row = sales_by_crop_map.get(row.id)
        crop_performance.append(
            {
                "crop_id": row.id,
                "crop_name": row.name,
                "category": row.category,
                "produced_qty": row.qty,
                "production_entries": row.entries,
                "sold_qty": sales_row.sold_qty if sales_row else 0,
                "revenue": sales_row.revenue if sales_row else 0,
            }
        )

    top_consumption_by_crop = (
        consumptions_query.join(Crop, Crop.id == CropConsumption.crop_id)
        .with_entities(
            Crop.id,
            Crop.name,
            func.coalesce(func.sum(CropConsumption.quantity_used), 0).label("used_qty"),
            func.count(CropConsumption.id).label("entries"),
        )
        .group_by(Crop.id, Crop.name)
        .order_by(desc("used_qty"))
        .limit(20)
        .all()
    )

    recent_productions = production_query.order_by(Production.production_date.desc()).limit(40).all()
    recent_crop_consumptions = (
        consumptions_query.order_by(CropConsumption.consumption_date.desc()).limit(40).all()
    )

    export_sections = [
        {
            "title": "Ù…Ù„Ø®Øµ Ø§Ù„Ø¥Ù†ØªØ§Ø¬",
            "headers": ["Ø§Ù„Ù…Ø¤Ø´Ø±", "Ø§Ù„Ù‚ÙŠÙ…Ø©"],
            "rows": [
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ø£ØµÙ†Ø§Ù", len(crops)],
                ["Ø§Ù„Ø£ØµÙ†Ø§Ù Ø§Ù„Ù†Ø´Ø·Ø©", active_crops],
                ["Ø³Ø¬Ù„Ø§Øª Ø§Ù„Ø¥Ù†ØªØ§Ø¬ (Ø¶Ù…Ù† Ø§Ù„ÙÙ„ØªØ±Ø©)", production_records_count],
                ["ÙƒÙ…ÙŠØ© Ø§Ù„Ø¥Ù†ØªØ§Ø¬ (Ø¶Ù…Ù† Ø§Ù„ÙÙ„ØªØ±Ø©)", total_production_qty],
                ["Ø³Ø¬Ù„Ø§Øª Ø§Ø³ØªÙ‡Ù„Ø§Ùƒ Ø§Ù„Ø£ØµÙ†Ø§Ù (Ø¶Ù…Ù† Ø§Ù„ÙÙ„ØªØ±Ø©)", crop_consumption_count],
                ["ÙƒÙ…ÙŠØ© Ø§Ù„Ø§Ø³ØªÙ‡Ù„Ø§Ùƒ (Ø¶Ù…Ù† Ø§Ù„ÙÙ„ØªØ±Ø©)", crop_consumption_qty],
            ],
        },
        {
            "title": "Ø£Ø¯Ø§Ø¡ Ø§Ù„Ø£ØµÙ†Ø§Ù",
            "headers": ["Ø§Ù„ØµÙ†Ù", "Ø§Ù„ÙØ¦Ø©", "ÙƒÙ…ÙŠØ© Ø§Ù„Ø¥Ù†ØªØ§Ø¬", "Ø³Ø¬Ù„Ø§Øª Ø§Ù„Ø¥Ù†ØªØ§Ø¬", "ÙƒÙ…ÙŠØ© Ø§Ù„Ù…Ø¨ÙŠØ¹Ø§Øª", "Ø§Ù„Ø¥ÙŠØ±Ø§Ø¯"],
            "rows": [
                [
                    row["crop_name"],
                    row["category"],
                    row["produced_qty"],
                    row["production_entries"],
                    row["sold_qty"],
                    row["revenue"],
                ]
                for row in crop_performance
            ],
        },
        {
            "title": "Ø£Ø¹Ù„Ù‰ Ø§Ø³ØªÙ‡Ù„Ø§Ùƒ Ù„Ù„Ø£ØµÙ†Ø§Ù",
            "headers": ["Ø§Ù„ØµÙ†Ù", "ÙƒÙ…ÙŠØ© Ø§Ù„Ø§Ø³ØªÙ‡Ù„Ø§Ùƒ", "Ø¹Ø¯Ø¯ Ø§Ù„Ø³Ø¬Ù„Ø§Øª"],
            "rows": [[row.name, row.used_qty, row.entries] for row in top_consumption_by_crop],
        },
        {
            "title": "Ø¢Ø®Ø± Ø³Ø¬Ù„Ø§Øª Ø§Ù„Ø¥Ù†ØªØ§Ø¬",
            "headers": ["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„ØµÙ†Ù", "Ø§Ù„ÙƒÙ…ÙŠØ©", "Ø§Ù„ÙˆØ­Ø¯Ø©", "Ø§Ù„Ø¬ÙˆØ¯Ø©"],
            "rows": [
                [production.production_date, production.crop.name, production.quantity, production.unit, production.quality or "-"]
                for production in recent_productions
            ],
        },
        {
            "title": "Ø¢Ø®Ø± Ø§Ø³ØªÙ‡Ù„Ø§Ùƒ Ø§Ù„Ø£ØµÙ†Ø§Ù",
            "headers": ["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„ØµÙ†Ù", "Ø§Ù„Ù…Ø§Ø¯Ø©", "Ø§Ù„ÙƒÙ…ÙŠØ©"],
            "rows": [
                [
                    consumption.consumption_date,
                    consumption.crop.name,
                    consumption.inventory_item.name,
                    consumption.quantity_used,
                ]
                for consumption in recent_crop_consumptions
            ],
        },
    ]
    export_response = _maybe_export(
        "production_detailed_report",
        "ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø¥Ù†ØªØ§Ø¬ Ø§Ù„ØªÙØµÙŠÙ„ÙŠ",
        export_sections,
        from_date,
        to_date,
    )
    if export_response:
        return export_response

    return render_template(
        "reports/production_detailed_report.html",
        crops=crops,
        active_crops=active_crops,
        total_production_qty=total_production_qty,
        production_records_count=production_records_count,
        crop_consumption_qty=crop_consumption_qty,
        crop_consumption_count=crop_consumption_count,
        crop_performance=crop_performance,
        top_consumption_by_crop=top_consumption_by_crop,
        recent_productions=recent_productions,
        recent_crop_consumptions=recent_crop_consumptions,
        from_date=from_date_str,
        to_date=to_date_str,
        date_range_label=_date_range_label(from_date, to_date),
    )


@bp.route("/sales-detailed")
@login_required
def sales_detailed_report():
    """Detailed sales report."""
    denied = _require_reports_access()
    if denied:
        return denied

    from_date, to_date, from_date_str, to_date_str = _read_date_range()

    sales_query = _apply_date_range(Sales.query, Sales.sale_date, from_date, to_date)
    sales_records = sales_query.order_by(Sales.sale_date.desc()).all()
    sales_count = len(sales_records)
    total_quantity = sum(record.quantity or 0 for record in sales_records)
    total_revenue = sum(record.total_price or 0 for record in sales_records)
    paid_revenue = sum(
        (record.total_price or 0)
        for record in sales_records
        if record.payment_status == "Ù…Ø¯ÙÙˆØ¹"
    )
    unpaid_revenue = total_revenue - paid_revenue

    sales_by_crop = (
        sales_query.join(Crop, Crop.id == Sales.crop_id)
        .with_entities(
            Crop.name.label("crop_name"),
            func.count(Sales.id).label("entries"),
            func.coalesce(func.sum(Sales.quantity), 0).label("qty"),
            func.coalesce(func.sum(Sales.total_price), 0).label("revenue"),
        )
        .group_by(Crop.name)
        .order_by(desc("revenue"))
        .all()
    )

    sales_by_buyer = (
        sales_query.with_entities(
            Sales.buyer_name,
            func.count(Sales.id).label("entries"),
            func.coalesce(func.sum(Sales.total_price), 0).label("revenue"),
            func.coalesce(func.sum(Sales.quantity), 0).label("qty"),
        )
        .filter(Sales.buyer_name.isnot(None), Sales.buyer_name != "")
        .group_by(Sales.buyer_name)
        .order_by(desc("revenue"))
        .all()
    )

    payment_status_summary = (
        sales_query.with_entities(
            Sales.payment_status,
            func.count(Sales.id).label("entries"),
            func.coalesce(func.sum(Sales.total_price), 0).label("amount"),
        )
        .group_by(Sales.payment_status)
        .all()
    )

    recent_unpaid = (
        sales_query.filter(Sales.payment_status != "Ù…Ø¯ÙÙˆØ¹")
        .order_by(Sales.sale_date.desc())
        .limit(25)
        .all()
    )

    export_sections = [
        {
            "title": "Ù…Ù„Ø®Øµ Ø§Ù„Ù…Ø¨ÙŠØ¹Ø§Øª",
            "headers": ["Ø§Ù„Ù…Ø¤Ø´Ø±", "Ø§Ù„Ù‚ÙŠÙ…Ø©"],
            "rows": [
                ["Ø¹Ø¯Ø¯ Ø¹Ù…Ù„ÙŠØ§Øª Ø§Ù„Ø¨ÙŠØ¹", sales_count],
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„ÙƒÙ…ÙŠØ©", total_quantity],
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ø¥ÙŠØ±Ø§Ø¯", total_revenue],
                ["Ø¥ÙŠØ±Ø§Ø¯ Ù…Ø­ØµÙ„", paid_revenue],
                ["Ø¥ÙŠØ±Ø§Ø¯ ØºÙŠØ± Ù…Ø­ØµÙ„", unpaid_revenue],
            ],
        },
        {
            "title": "Ø§Ù„ØªØ­ØµÙŠÙ„ Ø­Ø³Ø¨ Ø§Ù„Ø­Ø§Ù„Ø©",
            "headers": ["Ø§Ù„Ø­Ø§Ù„Ø©", "Ø¹Ø¯Ø¯ Ø§Ù„Ø³Ø¬Ù„Ø§Øª", "Ø§Ù„Ù‚ÙŠÙ…Ø©"],
            "rows": [[row.payment_status, row.entries, row.amount] for row in payment_status_summary],
        },
        {
            "title": "Ø£Ø¹Ù„Ù‰ Ø§Ù„Ù…Ø´ØªØ±ÙŠÙ†",
            "headers": ["Ø§Ù„Ù…Ø´ØªØ±ÙŠ", "Ø¹Ø¯Ø¯ Ø§Ù„Ø¹Ù…Ù„ÙŠØ§Øª", "Ø§Ù„ÙƒÙ…ÙŠØ©", "Ø§Ù„Ø¥ÙŠØ±Ø§Ø¯"],
            "rows": [[row.buyer_name, row.entries, row.qty, row.revenue] for row in sales_by_buyer],
        },
        {
            "title": "Ø§Ù„Ù…Ø¨ÙŠØ¹Ø§Øª Ø­Ø³Ø¨ Ø§Ù„ØµÙ†Ù",
            "headers": ["Ø§Ù„ØµÙ†Ù", "Ø¹Ø¯Ø¯ Ø§Ù„Ø¹Ù…Ù„ÙŠØ§Øª", "Ø§Ù„ÙƒÙ…ÙŠØ©", "Ø§Ù„Ø¥ÙŠØ±Ø§Ø¯"],
            "rows": [[row.crop_name, row.entries, row.qty, row.revenue] for row in sales_by_crop],
        },
        {
            "title": "Ø¢Ø®Ø± Ø§Ù„Ù…Ø¨ÙŠØ¹Ø§Øª ØºÙŠØ± Ø§Ù„Ù…Ø¯ÙÙˆØ¹Ø©",
            "headers": ["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„ØµÙ†Ù", "Ø§Ù„Ù…Ø´ØªØ±ÙŠ", "Ø§Ù„Ù‚ÙŠÙ…Ø©", "Ø§Ù„Ø­Ø§Ù„Ø©"],
            "rows": [
                [
                    row.sale_date,
                    row.crop.name,
                    row.buyer_name or "-",
                    row.total_price,
                    row.payment_status,
                ]
                for row in recent_unpaid
            ],
        },
    ]
    export_response = _maybe_export(
        "sales_detailed_report",
        "ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ù…Ø¨ÙŠØ¹Ø§Øª Ø§Ù„ØªÙØµÙŠÙ„ÙŠ",
        export_sections,
        from_date,
        to_date,
    )
    if export_response:
        return export_response

    return render_template(
        "reports/sales_detailed_report.html",
        sales_records=sales_records,
        sales_count=sales_count,
        total_quantity=total_quantity,
        total_revenue=total_revenue,
        paid_revenue=paid_revenue,
        unpaid_revenue=unpaid_revenue,
        sales_by_crop=sales_by_crop,
        sales_by_buyer=sales_by_buyer,
        payment_status_summary=payment_status_summary,
        recent_unpaid=recent_unpaid,
        from_date=from_date_str,
        to_date=to_date_str,
        date_range_label=_date_range_label(from_date, to_date),
    )


@bp.route("/accounting-detailed")
@login_required
def accounting_detailed_report():
    """Detailed accounting report."""
    denied = _require_financial_access()
    if denied:
        return denied

    from_date, to_date, from_date_str, to_date_str = _read_date_range()

    transactions_query = _apply_date_range(
        Transaction.query, Transaction.transaction_date, from_date, to_date
    )
    transactions = transactions_query.order_by(Transaction.transaction_date.desc()).all()
    for transaction in transactions:
        transaction.display_transaction_type = normalize_transaction_type(transaction.transaction_type)
        transaction.display_reference_type = normalize_reference_type(transaction.reference_type)

    total_income = sum(t.amount or 0 for t in transactions if is_income_transaction(t.transaction_type))
    total_expenses = sum(t.amount or 0 for t in transactions if is_expense_transaction(t.transaction_type))
    income_count = sum(1 for t in transactions if is_income_transaction(t.transaction_type))
    expense_count = sum(1 for t in transactions if is_expense_transaction(t.transaction_type))
    net_profit = total_income - total_expenses

    expense_by_category = (
        transactions_query.join(ExpenseCategory, ExpenseCategory.id == Transaction.category_id)
        .with_entities(
            ExpenseCategory.name,
            func.coalesce(func.sum(Transaction.amount), 0).label("amount"),
            func.count(Transaction.id).label("entries"),
        )
        .filter(Transaction.transaction_type.in_(EXPENSE_TRANSACTION_TYPE_ALIASES))
        .group_by(ExpenseCategory.name)
        .order_by(desc("amount"))
        .all()
    )

    expense_by_reference = (
        transactions_query.with_entities(
            Transaction.reference_type,
            func.coalesce(func.sum(Transaction.amount), 0).label("amount"),
            func.count(Transaction.id).label("entries"),
        )
        .filter(Transaction.transaction_type.in_(EXPENSE_TRANSACTION_TYPE_ALIASES))
        .group_by(Transaction.reference_type)
        .order_by(desc("amount"))
        .all()
    )

    monthly_summary = (
        transactions_query.with_entities(
            func.strftime("%Y-%m", Transaction.transaction_date).label("month"),
            func.sum(
                case(
                    (Transaction.transaction_type.in_(INCOME_TRANSACTION_TYPE_ALIASES), Transaction.amount),
                    else_=0,
                )
            ).label("income"),
            func.sum(
                case(
                    (Transaction.transaction_type.in_(EXPENSE_TRANSACTION_TYPE_ALIASES), Transaction.amount),
                    else_=0,
                )
            ).label("expense"),
        )
        .group_by(func.strftime("%Y-%m", Transaction.transaction_date))
        .order_by(desc("month"))
        .limit(12)
        .all()
    )

    export_sections = [
        {
            "title": "Ù…Ù„Ø®Øµ Ø§Ù„Ù…Ø­Ø§Ø³Ø¨Ø©",
            "headers": ["Ø§Ù„Ù…Ø¤Ø´Ø±", "Ø§Ù„Ù‚ÙŠÙ…Ø©"],
            "rows": [
                ["Ø¹Ø¯Ø¯ Ø§Ù„Ù…Ø¹Ø§Ù…Ù„Ø§Øª", len(transactions)],
                ["Ø¹Ø¯Ø¯ Ù‚ÙŠÙˆØ¯ Ø§Ù„Ø¯Ø®Ù„", income_count],
                ["Ø¹Ø¯Ø¯ Ù‚ÙŠÙˆØ¯ Ø§Ù„Ù…ØµØ±ÙˆÙ", expense_count],
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ø¯Ø®Ù„", total_income],
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ù…ØµØ±ÙˆÙ", total_expenses],
                ["ØµØ§ÙÙŠ Ø§Ù„Ø±Ø¨Ø­", net_profit],
            ],
        },
        {
            "title": "Ø§Ù„Ù…ØµØ±ÙˆÙØ§Øª Ø­Ø³Ø¨ Ø§Ù„Ù…Ø±Ø¬Ø¹",
            "headers": ["Ø§Ù„Ù…Ø±Ø¬Ø¹", "Ø¹Ø¯Ø¯ Ø§Ù„Ø³Ø¬Ù„Ø§Øª", "Ø§Ù„Ù‚ÙŠÙ…Ø©"],
            "rows": [
                [row.reference_type or "ØºÙŠØ± Ù…Ø­Ø¯Ø¯", row.entries, row.amount]
                for row in expense_by_reference
            ],
        },
        {
            "title": "Ø§Ù„Ù…ØµØ±ÙˆÙØ§Øª Ø­Ø³Ø¨ Ø§Ù„ÙØ¦Ø©",
            "headers": ["Ø§Ù„ÙØ¦Ø©", "Ø¹Ø¯Ø¯ Ø§Ù„Ø³Ø¬Ù„Ø§Øª", "Ø§Ù„Ù‚ÙŠÙ…Ø©"],
            "rows": [[row.name, row.entries, row.amount] for row in expense_by_category],
        },
        {
            "title": "Ù…Ù„Ø®Øµ Ø´Ù‡Ø±ÙŠ",
            "headers": ["Ø§Ù„Ø´Ù‡Ø±", "Ø§Ù„Ø¯Ø®Ù„", "Ø§Ù„Ù…ØµØ±ÙˆÙ", "Ø§Ù„ØµØ§ÙÙŠ"],
            "rows": [
                [row.month, row.income or 0, row.expense or 0, (row.income or 0) - (row.expense or 0)]
                for row in monthly_summary
            ],
        },
        {
            "title": "Ø¢Ø®Ø± Ø§Ù„Ù…Ø¹Ø§Ù…Ù„Ø§Øª",
            "headers": ["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„Ù†ÙˆØ¹", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„ÙØ¦Ø©", "Ø§Ù„Ù‚ÙŠÙ…Ø©", "Ø§Ù„Ù…Ø±Ø¬Ø¹"],
            "rows": [
                [
                    transaction.transaction_date,
                    transaction.display_transaction_type,
                    transaction.description,
                    transaction.category.name if transaction.category else "-",
                    transaction.amount,
                    transaction.display_reference_type or "-",
                ]
                for transaction in transactions[:100]
            ],
        },
    ]
    export_response = _maybe_export(
        "accounting_detailed_report",
        "ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ù…Ø­Ø§Ø³Ø¨Ø© Ø§Ù„ØªÙØµÙŠÙ„ÙŠ",
        export_sections,
        from_date,
        to_date,
    )
    if export_response:
        return export_response

    return render_template(
        "reports/accounting_detailed_report.html",
        transactions=transactions,
        total_income=total_income,
        total_expenses=total_expenses,
        net_profit=net_profit,
        income_count=income_count,
        expense_count=expense_count,
        expense_by_category=expense_by_category,
        expense_by_reference=expense_by_reference,
        monthly_summary=monthly_summary,
        from_date=from_date_str,
        to_date=to_date_str,
        date_range_label=_date_range_label(from_date, to_date),
    )


@bp.route("/motors-detailed")
@login_required
def motors_detailed_report():
    """Detailed motors report."""
    denied = _require_reports_access()
    if denied:
        return denied

    from_date, to_date, from_date_str, to_date_str = _read_date_range()

    motors = Motor.query.order_by(Motor.is_active.desc(), Motor.name.asc()).all()
    active_motors = sum(1 for motor in motors if motor.is_active)
    inactive_motors = len(motors) - active_motors

    usage_query = _apply_date_range(MotorUsage.query, MotorUsage.usage_date, from_date, to_date)
    costs_query = _apply_date_range(MotorCost.query, MotorCost.cost_date, from_date, to_date)

    usage_count = usage_query.count()
    total_usage_hours = _query_value(usage_query, func.sum(MotorUsage.total_hours))
    total_fuel_added = _query_value(usage_query, func.sum(MotorUsage.fuel_added))
    total_fuel_cost = _query_value(usage_query, func.sum(MotorUsage.fuel_cost))

    cost_entries_count = costs_query.count()
    total_motor_cost = _query_value(costs_query, func.sum(MotorCost.total_cost))

    quotas = OperatorQuota.query.order_by(OperatorQuota.year.desc(), OperatorQuota.operator_name.asc()).all()
    allocated_quota_hours = sum(quota.allocated_hours or 0 for quota in quotas)
    used_quota_hours = sum(quota.used_hours or 0 for quota in quotas)
    remaining_quota_hours = sum(quota.remaining_hours or 0 for quota in quotas)

    top_motors_by_hours = (
        usage_query.join(Motor, Motor.id == MotorUsage.motor_id)
        .with_entities(
            Motor.name.label("motor_name"),
            func.coalesce(func.sum(MotorUsage.total_hours), 0).label("hours"),
            func.count(MotorUsage.id).label("entries"),
        )
        .group_by(Motor.name)
        .order_by(desc("hours"))
        .limit(15)
        .all()
    )

    top_operators_by_hours = (
        usage_query.with_entities(
            MotorUsage.operator_name,
            func.coalesce(func.sum(MotorUsage.total_hours), 0).label("hours"),
            func.count(MotorUsage.id).label("entries"),
            func.coalesce(func.sum(MotorUsage.fuel_cost), 0).label("fuel_cost"),
        )
        .group_by(MotorUsage.operator_name)
        .order_by(desc("hours"))
        .limit(20)
        .all()
    )

    recent_usage = usage_query.order_by(MotorUsage.usage_date.desc(), MotorUsage.created_at.desc()).limit(40).all()
    recent_costs = costs_query.order_by(MotorCost.cost_date.desc()).limit(40).all()
    motor_lookup = {motor.id: motor.name for motor in motors}

    export_sections = [
        {
            "title": "Ù…Ù„Ø®Øµ Ø§Ù„Ù…Ø­Ø±ÙƒØ§Øª",
            "headers": ["Ø§Ù„Ù…Ø¤Ø´Ø±", "Ø§Ù„Ù‚ÙŠÙ…Ø©"],
            "rows": [
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ù…Ø­Ø±ÙƒØ§Øª", len(motors)],
                ["Ù…Ø­Ø±ÙƒØ§Øª Ù†Ø´Ø·Ø©", active_motors],
                ["Ù…Ø­Ø±ÙƒØ§Øª Ù…Ø¹Ø·Ù„Ø©", inactive_motors],
                ["Ø³Ø¬Ù„Ø§Øª Ø§Ù„ØªØ´ØºÙŠÙ„ (Ø¶Ù…Ù† Ø§Ù„ÙÙ„ØªØ±Ø©)", usage_count],
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø³Ø§Ø¹Ø§Øª Ø§Ù„ØªØ´ØºÙŠÙ„", total_usage_hours],
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„ÙˆÙ‚ÙˆØ¯ Ø§Ù„Ù…Ø¶Ø§Ù", total_fuel_added],
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ ØªÙƒÙ„ÙØ© Ø§Ù„ÙˆÙ‚ÙˆØ¯", total_fuel_cost],
                ["Ø¹Ø¯Ø¯ Ø³Ø¬Ù„Ø§Øª Ø§Ù„ØªÙƒØ§Ù„ÙŠÙ (Ø¶Ù…Ù† Ø§Ù„ÙÙ„ØªØ±Ø©)", cost_entries_count],
                ["Ø¥Ø¬Ù…Ø§Ù„ÙŠ ØªÙƒØ§Ù„ÙŠÙ Ø§Ù„Ù…Ø­Ø±ÙƒØ§Øª", total_motor_cost],
                ["Ø³Ø§Ø¹Ø§Øª Ø­ØµØµ Ù…Ø®ØµØµØ©", allocated_quota_hours],
                ["Ø³Ø§Ø¹Ø§Øª Ø­ØµØµ Ù…Ø³ØªØ®Ø¯Ù…Ø©", used_quota_hours],
                ["Ø³Ø§Ø¹Ø§Øª Ø­ØµØµ Ù…ØªØ¨Ù‚ÙŠØ©", remaining_quota_hours],
            ],
        },
        {
            "title": "Ø£ÙØ¶Ù„ Ø§Ù„Ù…Ø­Ø±ÙƒØ§Øª Ø­Ø³Ø¨ Ø³Ø§Ø¹Ø§Øª Ø§Ù„ØªØ´ØºÙŠÙ„",
            "headers": ["Ø§Ù„Ù…Ø­Ø±Ùƒ", "Ø§Ù„Ø³Ø§Ø¹Ø§Øª", "Ø¹Ø¯Ø¯ Ø§Ù„Ø³Ø¬Ù„Ø§Øª"],
            "rows": [[row.motor_name, row.hours, row.entries] for row in top_motors_by_hours],
        },
        {
            "title": "Ø£ÙØ¶Ù„ Ø§Ù„Ù…Ø´ØºÙ„ÙŠÙ† Ø­Ø³Ø¨ Ø§Ù„Ø³Ø§Ø¹Ø§Øª",
            "headers": ["Ø§Ù„Ù…Ø´ØºÙ„", "Ø§Ù„Ø³Ø§Ø¹Ø§Øª", "Ø¹Ø¯Ø¯ Ø§Ù„Ø³Ø¬Ù„Ø§Øª", "ØªÙƒÙ„ÙØ© Ø§Ù„ÙˆÙ‚ÙˆØ¯"],
            "rows": [
                [row.operator_name, row.hours, row.entries, row.fuel_cost]
                for row in top_operators_by_hours
            ],
        },
        {
            "title": "Ø­ØµØµ Ø§Ù„Ù…Ø´ØºÙ„ÙŠÙ†",
            "headers": ["Ø§Ù„Ù…Ø´ØºÙ„", "Ø§Ù„Ø³Ù†Ø©", "Ø§Ù„Ù…Ø®ØµØµ", "Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…", "Ø§Ù„Ù…ØªØ¨Ù‚ÙŠ", "Ø§Ù„Ø­Ø§Ù„Ø©"],
            "rows": [
                [
                    quota.operator_name,
                    quota.year,
                    quota.allocated_hours,
                    quota.used_hours,
                    quota.remaining_hours,
                    quota.status,
                ]
                for quota in quotas
            ],
        },
        {
            "title": "Ø¢Ø®Ø± Ø³Ø¬Ù„Ø§Øª Ø§Ù„ØªØ´ØºÙŠÙ„",
            "headers": ["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„Ù…Ø­Ø±Ùƒ", "Ø§Ù„Ù…Ø´ØºÙ„", "Ø§Ù„Ø³Ø§Ø¹Ø§Øª"],
            "rows": [
                [usage.usage_date, usage.motor.name, usage.operator_name, usage.total_hours or 0]
                for usage in recent_usage
            ],
        },
        {
            "title": "Ø¢Ø®Ø± ØªÙƒØ§Ù„ÙŠÙ Ø§Ù„Ù…Ø­Ø±ÙƒØ§Øª",
            "headers": ["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„Ù…Ø­Ø±Ùƒ", "Ø§Ù„Ù†ÙˆØ¹", "Ø§Ù„ÙƒÙ…ÙŠØ©", "Ø§Ù„ØªÙƒÙ„ÙØ©"],
            "rows": [
                [
                    cost.cost_date,
                    motor_lookup.get(cost.motor_id, "-"),
                    cost.cost_type,
                    cost.quantity,
                    cost.total_cost,
                ]
                for cost in recent_costs
            ],
        },
    ]
    export_response = _maybe_export(
        "motors_detailed_report",
        "ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ù…Ø­Ø±ÙƒØ§Øª Ø§Ù„ØªÙØµÙŠÙ„ÙŠ",
        export_sections,
        from_date,
        to_date,
    )
    if export_response:
        return export_response

    return render_template(
        "reports/motors_detailed_report.html",
        motors=motors,
        active_motors=active_motors,
        inactive_motors=inactive_motors,
        usage_count=usage_count,
        total_usage_hours=total_usage_hours,
        total_fuel_added=total_fuel_added,
        total_fuel_cost=total_fuel_cost,
        cost_entries_count=cost_entries_count,
        total_motor_cost=total_motor_cost,
        quotas=quotas,
        allocated_quota_hours=allocated_quota_hours,
        used_quota_hours=used_quota_hours,
        remaining_quota_hours=remaining_quota_hours,
        top_motors_by_hours=top_motors_by_hours,
        top_operators_by_hours=top_operators_by_hours,
        recent_usage=recent_usage,
        recent_costs=recent_costs,
        motor_lookup=motor_lookup,
        from_date=from_date_str,
        to_date=to_date_str,
        date_range_label=_date_range_label(from_date, to_date),
    )


@bp.route("/crop/<int:crop_id>")
@login_required
def crop_report(crop_id):
    """Single crop report."""
    denied = _require_reports_access()
    if denied:
        return denied

    crop = Crop.query.get_or_404(crop_id)
    production_records = Production.query.filter_by(crop_id=crop_id).all()
    total_production = sum(record.quantity or 0 for record in production_records)

    sales_records = Sales.query.filter_by(crop_id=crop_id).all()
    total_sales = sum(record.quantity or 0 for record in sales_records)
    total_revenue = sum(record.total_price or 0 for record in sales_records)

    consumptions = crop.consumptions

    return render_template(
        "reports/crop_report.html",
        crop=crop,
        production_records=production_records,
        total_production=total_production,
        sales_records=sales_records,
        total_sales=total_sales,
        total_revenue=total_revenue,
        consumptions=consumptions,
    )


@bp.route("/worker/<int:worker_id>")
@login_required
def worker_report(worker_id):
    """Single worker report."""
    denied = _require_reports_access()
    if denied:
        return denied

    worker = Worker.query.get_or_404(worker_id)
    from_date, to_date, from_date_str, to_date_str = _read_date_range()

    work_logs_query = _apply_date_range(
        WorkLog.query.filter_by(worker_id=worker_id),
        WorkLog.work_date,
        from_date,
        to_date,
    )
    work_logs = work_logs_query.order_by(WorkLog.work_date.desc(), WorkLog.id.desc()).all()
    work_log_count = len(work_logs)

    attendance_query = _apply_date_range(
        Attendance.query.filter_by(worker_id=worker_id),
        Attendance.attendance_date,
        from_date,
        to_date,
    )
    attended_days = (
        attendance_query.filter(
            (Attendance.status == "حاضر") | (Attendance.is_present.is_(True))
        )
        .order_by(Attendance.attendance_date.desc(), Attendance.id.desc())
        .all()
    )

    if work_log_count:
        total_hours = sum(log.hours or 0 for log in work_logs)
        worked_days = len({log.work_date for log in work_logs})
        activity_rows = [
            {
                "activity_date": log.work_date,
                "activity_type": "ساعات عمل",
                "hours": float(log.hours or 0),
                "details": (log.shift_type or "-"),
                "location": (log.location or "-"),
                "notes": (log.notes or "-"),
            }
            for log in work_logs
        ]
    else:
        total_hours = sum(att.hours_worked or 0 for att in attended_days)
        worked_days = len({att.attendance_date for att in attended_days})
        activity_rows = [
            {
                "activity_date": att.attendance_date,
                "activity_type": f"حضور ({att.status or '-'})",
                "hours": float(att.hours_worked or 0),
                "details": (att.status or "-"),
                "location": "-",
                "notes": (att.notes or "-"),
            }
            for att in attended_days
        ]

    salary = (
        float(total_hours or 0) * float(worker.hourly_rate or 0)
        if not worker.is_monthly
        else float(worker.monthly_salary or 0)
    )

    worker_transactions_query = Transaction.query.filter(
        Transaction.reference_id == worker_id,
        Transaction.reference_type.in_(WORKER_REFERENCE_TYPE_ALIASES),
    )
    worker_transactions_query = _apply_date_range(
        worker_transactions_query,
        Transaction.transaction_date,
        from_date,
        to_date,
    )
    worker_transactions = (
        worker_transactions_query
        .order_by(Transaction.transaction_date.desc(), Transaction.id.desc())
        .all()
    )

    payment_rows = []
    total_worker_loans = 0.0
    total_worker_advances = 0.0
    total_worker_payments = 0.0

    for transaction in worker_transactions:
        if not is_expense_transaction(transaction.transaction_type):
            continue

        amount = float(transaction.amount or 0.0)
        kind = _detect_worker_payment_kind(transaction)
        if kind == "loan":
            total_worker_loans += amount
        elif kind == "advance":
            total_worker_advances += amount

        total_worker_payments += amount
        payment_rows.append(
            {
                "transaction_date": transaction.transaction_date,
                "description": transaction.description,
                "amount": amount,
                "kind": kind,
                "notes": transaction.notes or "-",
            }
        )

    remaining_salary = salary - total_worker_payments

    export_sections = [
        {
            "title": "ملخص العامل",
            "headers": ["المؤشر", "القيمة"],
            "rows": [
                ["العامل", worker.name],
                ["نوع العامل", "شهري" if worker.is_monthly else "بالساعة"],
                ["إجمالي ساعات العمل", total_hours],
                ["عدد أيام العمل", worked_days],
                ["الأجر الأساسي", salary],
                ["إجمالي السلف", total_worker_loans],
                ["إجمالي الدفعات على الحساب", total_worker_advances],
                ["إجمالي الحسومات على العامل", total_worker_payments],
                ["المتبقي للعامل", remaining_salary],
            ],
        },
        {
            "title": "سجل العمل",
            "headers": ["التاريخ", "النوع", "الساعات", "التفاصيل", "الموقع", "الملاحظات"],
            "rows": [
                [
                    row["activity_date"],
                    row["activity_type"],
                    row["hours"],
                    row["details"],
                    row["location"],
                    row["notes"],
                ]
                for row in activity_rows
            ],
        },
        {
            "title": "سجل السلف والدفعات",
            "headers": ["التاريخ", "البيان", "النوع", "القيمة", "الملاحظات"],
            "rows": [
                [
                    row["transaction_date"],
                    row["description"],
                    "سلفة"
                    if row["kind"] == "loan"
                    else "دفعة على الحساب"
                    if row["kind"] == "advance"
                    else "قيد مرتبط بالعامل",
                    row["amount"],
                    row["notes"],
                ]
                for row in payment_rows
            ],
        },
    ]
    export_response = _maybe_export(
        "worker_report",
        f"تقرير العامل - {worker.name}",
        export_sections,
        from_date,
        to_date,
    )
    if export_response:
        return export_response

    return render_template(
        "reports/worker_report.html",
        worker=worker,
        activity_rows=activity_rows,
        payment_rows=payment_rows,
        total_hours=total_hours,
        worked_days=worked_days,
        salary=salary,
        total_worker_loans=total_worker_loans,
        total_worker_advances=total_worker_advances,
        total_worker_payments=total_worker_payments,
        remaining_salary=remaining_salary,
        from_date=from_date_str,
        to_date=to_date_str,
        date_range_label=_date_range_label(from_date, to_date),
    )


@bp.route("/financial")
@login_required
def financial_report():
    """Financial report."""
    denied = _require_financial_access()
    if denied:
        return denied

    transactions = Transaction.query.all()
    for transaction in transactions:
        transaction.display_transaction_type = normalize_transaction_type(transaction.transaction_type)

    total_income = sum(t.amount or 0 for t in transactions if is_income_transaction(t.transaction_type))
    total_expenses = sum(t.amount or 0 for t in transactions if is_expense_transaction(t.transaction_type))
    income_count = sum(1 for t in transactions if is_income_transaction(t.transaction_type))
    expense_count = sum(1 for t in transactions if is_expense_transaction(t.transaction_type))

    by_category = {}
    for transaction in transactions:
        if is_expense_transaction(transaction.transaction_type) and transaction.category:
            category_name = transaction.category.name
            by_category[category_name] = by_category.get(category_name, 0) + (transaction.amount or 0)

    return render_template(
        "reports/financial_report.html",
        transactions=transactions,
        total_income=total_income,
        total_expenses=total_expenses,
        income_count=income_count,
        expense_count=expense_count,
        by_category=by_category,
        net=total_income - total_expenses,
    )


@bp.route("/monthly")
@login_required
def monthly_report():
    """Monthly report."""
    denied = _require_reports_access()
    if denied:
        return denied

    now = datetime.now()
    month_start = datetime(now.year, now.month, 1)
    month_end = month_start + timedelta(days=32)
    month_end = month_end.replace(day=1) - timedelta(days=1)

    sales = Sales.query.filter(
        Sales.sale_date >= month_start.date(),
        Sales.sale_date <= month_end.date(),
    ).all()
    total_sales = sum(sale.total_price or 0 for sale in sales)

    production = Production.query.filter(
        Production.production_date >= month_start.date(),
        Production.production_date <= month_end.date(),
    ).all()

    expenses = Transaction.query.filter(
        Transaction.transaction_type.in_(EXPENSE_TRANSACTION_TYPE_ALIASES),
        Transaction.transaction_date >= month_start.date(),
        Transaction.transaction_date <= month_end.date(),
    ).all()
    total_expenses = sum(expense.amount or 0 for expense in expenses)

    return render_template(
        "reports/monthly_report.html",
        sales=sales,
        total_sales=total_sales,
        production=production,
        expenses=expenses,
        total_expenses=total_expenses,
        month=month_start.strftime("%B %Y"),
    )


def db_value(aggregate_expression, model):
    """Return numeric aggregate value with 0 fallback."""
    return model.query.with_entities(func.coalesce(aggregate_expression, 0)).scalar() or 0

